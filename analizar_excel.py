import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

def analizar_excel():
    ruta_excel = "Templates/TABLA A - CONVENIO.xlsm"
    
    try:
        # Cargar el workbook
        wb = load_workbook(ruta_excel, read_only=True, keep_vba=True)
        
        print("=== ANÁLISIS DEL ARCHIVO EXCEL ===")
        print(f"Nombre: {os.path.basename(ruta_excel)}")
        print(f"Hojas: {wb.sheetnames}")
        print(f"¿Contiene macros?: {wb.vba}")
        
        # Analizar cada hoja
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- HOJA: {sheet_name} ---")
            print(f"Dimensiones: {ws.dimensions}")
            print(f"Max fila: {ws.max_row}, Max columna: {ws.max_column}")
            
            # Buscar botones o controles
            print("Buscando botones/controles...")
            for drawing in ws._drawings:
                print(f"  - Control encontrado: {type(drawing).__name__}")
            
            # Buscar celdas con fórmulas (especialmente sumatorias)
            print("Buscando celdas con fórmulas...")
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # fórmula
                        print(f"  - Fórmula en {cell.coordinate}: {cell.value}")
            
            # Buscar celdas con valores (estructura de la tabla)
            print("Estructura de datos principales:")
            for row in range(1, min(20, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(10, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        row_data.append(f"{cell.value}")
                    else:
                        row_data.append("")
                if any(row_data):  # Si la fila tiene datos
                    print(f"  Fila {row}: {row_data[:5]}")  # Primeras 5 columnas
        
        wb.close()
        
    except Exception as e:
        print(f"Error al analizar el Excel: {e}")

if __name__ == "__main__":
    analizar_excel()
