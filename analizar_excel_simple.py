import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

def analizar_excel():
    ruta_excel = "Templates/TABLA A - CONVENIO.xlsm"
    
    try:
        # Cargar el workbook
        wb = load_workbook(ruta_excel, read_only=True)
        
        print("=== ANÁLISIS DEL ARCHIVO EXCEL ===")
        print(f"Nombre: {os.path.basename(ruta_excel)}")
        print(f"Hojas: {wb.sheetnames}")
        
        # Analizar la hoja principal
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        print(f"\n--- HOJA: {sheet_name} ---")
        print(f"Dimensiones: {ws.dimensions}")
        print(f"Max fila: {ws.max_row}, Max columna: {ws.max_column}")
        
        # Buscar botones o controles
        print("\nBuscando controles...")
        try:
            if hasattr(ws, '_drawings') and ws._drawings:
                for drawing in ws._drawings:
                    print(f"  - Control: {type(drawing).__name__}")
            else:
                print("  - No se encontraron controles visibles")
        except:
            print("  - No se pudieron analizar controles")
        
        # Buscar celdas con fórmulas (especialmente sumatorias)
        print("\nBuscando celdas con fórmulas...")
        formulas_found = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # fórmula
                    print(f"  - Fórmula en {cell.coordinate}: {cell.value}")
                    formulas_found += 1
        if formulas_found == 0:
            print("  - No se encontraron fórmulas")
        
        # Mostrar estructura de la tabla (primeras 50 filas)
        print("\nEstructura de datos:")
        for row in range(1, min(51, ws.max_row + 1)):
            row_data = []
            has_data = False
            for col in range(1, min(15, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip():
                    row_data.append(str(cell.value)[:20])  # Limitar longitud
                    has_data = True
                else:
                    row_data.append("")
            if has_data:
                print(f"  Fila {row:2d}: {' | '.join(row_data[:8])}")  # Primeras 8 columnas
        
        # Buscar patrones de sumatorias o totales
        print("\nBuscando posibles celdas de total/sumatoria...")
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    val_upper = cell.value.upper()
                    if any(keyword in val_upper for keyword in ['TOTAL', 'SUMA', 'SUMATORIA']):
                        print(f"  - Posible total en {cell.coordinate}: '{cell.value}'")
        
        wb.close()
        
    except Exception as e:
        print(f"Error al analizar el Excel: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analizar_excel()
